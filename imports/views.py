import json
from datetime import datetime

import pandas as pd
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone

from leads.models import Lead, SourceCategory, LeadSource, Course, LeadStage
from followups.models import FollowUp, FollowUpStatus, FollowUpMode
from .models import ImportJob, ImportError as ImportErrorModel
from . import cleaning

TARGET_FIELDS = [
    ("IGNORE", "— Ignore this column (ID / Ad ID / Non-required) —"),
    ("name", "Patient / Lead Name"),
    ("mobile", "Mobile / Phone Number"),
    ("email", "Email Address"),
    ("gender", "Gender (Male / Female / Other)"),
    ("age", "Age"),
    ("city", "City / Location / Area"),
    ("doctor", "Doctor / Consultant"),
    ("department", "Department / Speciality"),
    ("campaign", "Campaign Name"),
    ("source", "Lead Source / Platform (FB, IG, Google, etc.)"),
    ("assigned_to", "Assigned To / Telecaller / Executive"),
    ("inquiry_date", "Date / Created At (DD/MM/YYYY)"),
    ("notes", "Comments / Notes / Survey Question Responses"),
]

GUESS_KEYWORDS = {
    "name": ["patient name", "patient", "full name", "customer name", "lead name", "name"],
    "mobile": ["mobile number", "mobile", "phone number", "phone", "contact number", "contact", "call number", "whatsapp number", "cell"],
    "email": ["email", "e-mail", "mail"],
    "gender": ["gender", "sex", "m/f"],
    "age": ["age", "years", "yrs"],
    "city": ["city", "location", "address", "area", "town", "district"],
    "doctor": ["doctor", "dr name", "consultant", "physician", "surgeon"],
    "department": ["department", "speciality", "dept", "specialization"],
    "campaign": ["campaign name", "campaign", "ad name", "ad set name"],
    "source": ["origin", "source", "platform", "publisher platform", "lead source", "channel"],
    "assigned_to": ["assigned to", "assigned", "telecaller", "executive", "attendant", "caller", "agent", "lead owner", "owner", "assignee", "counsellor"],
    "inquiry_date": ["created at", "created_at", "date", "created time", "lead date", "inquiry date", "lead time"],
    "notes": ["remark", "comment", "issue", "note", "problem", "symptom", "query", "reason", "question"],
}


def _resolve_assigned_user(raw_val, hospital=None, default_user=None):
    """Finds or matches a User/Telecaller by name or username, or returns default_user."""
    if not raw_val or str(raw_val).strip() in ("", "-", "nan", "NaT", "none", "null"):
        return default_user
    from accounts.models import User
    from django.db.models import Q
    s = str(raw_val).strip()
    s_low = s.lower()
    
    qs = User.objects.filter(is_active=True)
    if hospital:
        qs_h = qs.filter(hospital=hospital)
        if qs_h.exists():
            qs = qs_h

    # 1. Exact username or full name match
    for u in qs:
        uname = (u.username or "").strip().lower()
        fname = (u.get_full_name() or "").strip().lower()
        if s_low == uname or s_low == fname or (fname and s_low in fname) or (uname and s_low in uname):
            return u
            
    # 2. First name / Last name partial match
    for u in qs:
        if (u.first_name and u.first_name.lower() in s_low) or (u.last_name and u.last_name.lower() in s_low):
            return u

    return default_user


def _guess_field(header):
    h = str(header).strip().lower()
    for field, keywords in GUESS_KEYWORDS.items():
        if any(k in h for k in keywords):
            return field
    return "IGNORE"


def _is_date_header(header):
    if isinstance(header, (datetime,)):
        return True
    try:
        import datetime as dt
        if isinstance(header, dt.date):
            return True
    except Exception:
        pass
    return cleaning.parse_date(header) is not None and len(str(header)) >= 6


def _detect_header_row(raw_df):
    for i in range(min(5, len(raw_df))):
        row = raw_df.iloc[i]
        non_null = row.notna().sum()
        text_like = sum(1 for v in row if isinstance(v, str) and len(v.strip()) > 1)
        if non_null >= 3 and text_like >= 2:
            return i
    return 0


def _get_or_create_source(category_name, source_name):
    if not category_name or not source_name:
        return None, None
    cat, _ = SourceCategory.objects.get_or_create(name=category_name)
    src, _ = LeadSource.objects.get_or_create(name=source_name, category=cat)
    return cat, src


def _get_or_create_course(name):
    if not name:
        return None
    course, _ = Course.objects.get_or_create(name=name)
    return course


def _default_stage():
    stage = LeadStage.objects.order_by("order").first()
    if not stage:
        stage = LeadStage.objects.create(name="New", order=0)
    return stage


def _load_excel_or_csv(file_path, filename=""):
    """Robustly reads Excel (.xlsx, .xls, .xlsm, .xlsb) or CSV with automatic engine fallback and multi-encoding support."""
    fn_lower = filename.lower()
    
    # 1. Try CSV parsing if filename ends with .csv or fallback
    if fn_lower.endswith(".csv"):
        for enc in ["utf-8-sig", "utf-8", "latin1", "cp1252", "iso-8859-1"]:
            try:
                df = pd.read_csv(file_path, encoding=enc)
                return {"type": "df", "df": df, "sheets": ["CSV Data"]}
            except Exception:
                continue
                
    # 2. Try pd.ExcelFile with automatic and explicit engine fallbacks
    for eng in [None, "openpyxl", "xlrd", "pyxlsb"]:
        try:
            if eng:
                xl = pd.ExcelFile(file_path, engine=eng)
            else:
                xl = pd.ExcelFile(file_path)
            return {"type": "excel", "xl": xl, "sheets": xl.sheet_names, "engine": eng}
        except Exception:
            continue
            
    # 3. Last fallback: Try reading as CSV even if named .xlsx/.xls
    for enc in ["utf-8-sig", "utf-8", "latin1", "cp1252"]:
        try:
            df = pd.read_csv(file_path, encoding=enc)
            return {"type": "df", "df": df, "sheets": ["Imported Data"]}
        except Exception:
            continue
            
    raise ValueError("File format could not be read. Please upload a valid .xlsx, .xls, or .csv file.")


from django.core.cache import cache
from accounts.models import User, Hospital
from leads.models import Campaign as HospitalCampaign
from .nel_hospital_import_service import (
    parse_any_file_to_dataframe,
    extract_campaign_lead_data,
    check_duplicates_in_db,
    generate_lead_code,
    parse_flexible_date,
)
from django.db import transaction


def _can_user_access_import(user):
    """
    Lead Attendant, Hospital Manager, Hospital Admin, Super Admin can import campaign leads.
    """
    if not user.is_authenticated:
        return False
    if user.is_superuser or user.role in (User.Role.SUPER_ADMIN, User.Role.ADMIN, User.Role.MANAGER, User.Role.LEAD_ATTENDENT):
        return True
    return getattr(user, "can_import_export", False)


@login_required
@user_passes_test(_can_user_access_import)
def upload(request):
    """
    Unified Import Leads View:
    - Mode 1: Campaign Leads (Available to Lead Attendant, Manager, Admin, Super Admin)
    - Mode 2: Previous Lead Data (Available to Manager, Admin, Super Admin)
    - Zappcode Super Admin can select target hospital/business.
    """
    user = request.user
    is_super_admin_no_hospital = bool(user.role == User.Role.SUPER_ADMIN and not user.hospital)
    can_import_previous = bool(user.is_superuser or user.role in (User.Role.SUPER_ADMIN, User.Role.ADMIN, User.Role.MANAGER))
    
    # Available campaigns with current leads count
    from django.db.models import Count
    if user.hospital:
        campaigns = HospitalCampaign.objects.filter(hospital=user.hospital, is_active=True).annotate(leads_count=Count("leads")).order_by("-id")
        current_leads_count = Lead.objects.filter(hospital=user.hospital, is_archived=False).count()
    else:
        campaigns = HospitalCampaign.objects.filter(is_active=True).annotate(leads_count=Count("leads")).order_by("-id")
        current_leads_count = Lead.objects.filter(is_archived=False).count()

    all_hospitals = []
    if is_super_admin_no_hospital:
        all_hospitals = Hospital.objects.filter(is_active=True).annotate(leads_count=Count("leads")).order_by("name")

    context = {
        "active": "import",
        "is_super_admin_no_hospital": is_super_admin_no_hospital,
        "can_import_previous": can_import_previous,
        "available_campaigns": campaigns,
        "all_hospitals": all_hospitals,
        "current_leads_count": current_leads_count,
    }
    return render(request, "imports/upload.html", context)


@login_required
@user_passes_test(_can_user_access_import)
def ajax_create_campaign(request):
    """Creates a new Campaign via AJAX from the import screen."""
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Invalid request method"}, status=400)
    
    name = request.POST.get("name", "").strip()
    platform = request.POST.get("platform", "Meta Ads").strip()
    ad_set = request.POST.get("ad_set", "").strip()
    hospital_id = request.POST.get("hospital_id", "").strip()

    if not name:
        return JsonResponse({"success": False, "error": "Campaign name is required."})

    target_hospital = None
    if hospital_id:
        target_hospital = Hospital.objects.filter(pk=hospital_id).first()
    elif request.user.hospital:
        target_hospital = request.user.hospital

    start_date_str = request.POST.get("start_date", "").strip()
    end_date_str = request.POST.get("end_date", "").strip()

    start_date = parse_flexible_date(start_date_str) if start_date_str else None
    end_date = parse_flexible_date(end_date_str) if end_date_str else None

    defaults = {
        "platform": platform,
        "ad_set": ad_set,
        "is_active": True,
    }
    if start_date:
        defaults["start_date"] = start_date
    if end_date:
        defaults["end_date"] = end_date

    campaign, created = HospitalCampaign.objects.get_or_create(
        name=name,
        hospital=target_hospital,
        defaults=defaults
    )
    if not created:
        if start_date and not campaign.start_date:
            campaign.start_date = start_date
        if end_date and not campaign.end_date:
            campaign.end_date = end_date
        campaign.save()

    return JsonResponse({
        "success": True,
        "campaign": {
            "id": campaign.id,
            "name": campaign.name,
            "platform": campaign.platform,
            "hospital_id": campaign.hospital_id,
        }
    })


@login_required
@user_passes_test(_can_user_access_import)
def campaign_import_process(request):
    """
    Processes uploaded Campaign Lead file (.xml, .xlsx, .xls, .csv).
    If duplicate_strategy == 'preview', renders interactive conflict resolution screen.
    Otherwise executes directly.
    """
    if request.method != "POST" or "lead_file" not in request.FILES:
        messages.error(request, "Please select a valid leads file to upload.")
        return redirect("imports:upload")

    uploaded_file = request.FILES["lead_file"]
    campaign_id = request.POST.get("campaign_id")
    target_hospital_id = request.POST.get("target_hospital_id")
    duplicate_strategy = request.POST.get("duplicate_strategy", "preview")

    # Determine target hospital
    target_hospital = None
    if target_hospital_id:
        target_hospital = Hospital.objects.filter(pk=target_hospital_id).first()
    elif request.user.hospital:
        target_hospital = request.user.hospital

    # Determine Campaign
    campaign = None
    if campaign_id:
        campaign = HospitalCampaign.objects.filter(pk=campaign_id).first()

    if not campaign:
        messages.error(request, "Please select or create a Campaign for these leads.")
        return redirect("imports:upload")

    # Parse dataframe
    try:
        df = parse_any_file_to_dataframe(uploaded_file)
    except Exception as e:
        messages.error(request, f"Could not read the uploaded file: {e}")
        return redirect("imports:upload")

    if df is None or len(df) == 0:
        messages.error(request, "The uploaded file is empty.")
        return redirect("imports:upload")

    # Clean and extract campaign rows
    lead_rows = extract_campaign_lead_data(df, target_campaign=campaign, target_hospital=target_hospital)
    if not lead_rows:
        messages.error(request, "No valid lead rows could be extracted. Please check the file columns.")
        return redirect("imports:upload")

    # Check duplicates against DB
    processed_rows, stats = check_duplicates_in_db(lead_rows, hospital=target_hospital)

    # If preview strategy requested or duplicates exist
    if duplicate_strategy == "preview":
        import uuid
        cache_key = f"camp_import_{uuid.uuid4().hex}"
        cache_payload = {
            "rows": processed_rows,
            "campaign_id": campaign.id,
            "target_hospital_id": target_hospital.id if target_hospital else None,
            "original_filename": uploaded_file.name,
        }
        # Store in cache (file-based)
        cache.set(cache_key, cache_payload, timeout=86400)
        # Also store in DB session as fallback
        request.session[cache_key] = cache_payload
        request.session.modified = True

        context = {
            "active": "import",
            "campaign": campaign,
            "target_hospital": target_hospital,
            "original_filename": uploaded_file.name,
            "stats": stats,
            "rows": processed_rows,
            "cache_key": cache_key,
        }
        return render(request, "imports/nel_campaign_import_preview.html", context)

    # Automatic execution based on strategy (skip / create / update)
    return _execute_campaign_leads_import(
        request, processed_rows, campaign, target_hospital, uploaded_file.name, default_strategy=duplicate_strategy
    )


@login_required
@user_passes_test(_can_user_access_import)
def campaign_import_execute(request):
    """
    Executes the campaign leads import after user confirms duplicate actions.
    """
    if request.method != "POST":
        return redirect("imports:upload")

    cache_key = request.POST.get("cache_key")
    cached_data = cache.get(cache_key) if cache_key else None
    
    # Fallback to session if cache missed
    if not cached_data and cache_key and cache_key in request.session:
        cached_data = request.session.get(cache_key)

    if not cached_data:
        messages.error(request, "Import session expired or not found. Please upload the file again.")
        return redirect("imports:upload")

    rows = cached_data.get("rows", [])
    campaign_id = request.POST.get("campaign_id") or cached_data.get("campaign_id")
    target_hospital_id = request.POST.get("target_hospital_id") or cached_data.get("target_hospital_id")
    original_filename = cached_data.get("original_filename", "campaign_leads.xml")

    campaign = HospitalCampaign.objects.filter(pk=campaign_id).first()
    target_hospital = Hospital.objects.filter(pk=target_hospital_id).first() if target_hospital_id else request.user.hospital

    # Apply row-level action decisions from form
    for idx, r in enumerate(rows):
        action_val = request.POST.get(f"action_{idx}")
        if action_val:
            r["duplicate_action"] = action_val

    return _execute_campaign_leads_import(
        request, rows, campaign, target_hospital, original_filename
    )


def _execute_campaign_leads_import(request, rows, campaign, target_hospital, original_filename, default_strategy=None):
    """Core function to create/update Lead records with transaction safety and update Campaign start/end dates."""
    from leads.models import MasterGroup, MasterItem, LeadCustomField
    default_cat, _ = SourceCategory.objects.get_or_create(name="Digital Marketing", defaults={"order": 1})
    stage_new = LeadStage.objects.filter(name__iexact="New").first() or LeadStage.objects.first()

    job = ImportJob.objects.create(
        original_filename=original_filename,
        total_rows=len(rows),
        created_by=request.user,
        status=ImportJob.Status.PROCESSING,
    )

    imported_count = 0
    updated_count = 0
    skipped_count = 0

    source_cache = {}
    extracted_dates = []

    with transaction.atomic():
        for r in rows:
            action = r.get("duplicate_action") or default_strategy or "create"
            mobile = r.get("mobile", "")
            name = r.get("name", "Unknown Patient")
            email = r.get("email", "")
            inquiry_date = parse_flexible_date(r.get("inquiry_date"))
            if inquiry_date:
                extracted_dates.append(inquiry_date)
            source_name = r.get("source_name", "Instagram")
            notes = r.get("notes", "")
            custom_data = r.get("custom_data", {})
            raw_meta = r.get("raw_metadata", {})
            external_id = r.get("external_lead_id", "")

            # If duplicate and user chose to discard
            if r.get("is_duplicate") and action == "discard":
                skipped_count += 1
                continue

            # Lead Source resolution
            if source_name not in source_cache:
                src_obj, _ = LeadSource.objects.get_or_create(
                    name=source_name,
                    category=default_cat,
                    defaults={"is_active": True}
                )
                source_cache[source_name] = src_obj
            lead_source_obj = source_cache[source_name]

            # If duplicate and user chose update
            if r.get("is_duplicate") and action == "update" and r.get("existing_lead_id"):
                existing = Lead.objects.filter(pk=r["existing_lead_id"]).first()
                if existing:
                    if email and not existing.email:
                        existing.email = email
                    if campaign:
                        existing.campaign = campaign
                    if notes:
                        existing.notes = (existing.notes + "\n" + notes).strip()
                    if custom_data:
                        existing.custom_data.update(custom_data)
                    existing.import_job = job
                    existing.save()
                    updated_count += 1
                    continue

            # Create New Lead
            lead_code = generate_lead_code(hospital=target_hospital)
            
            new_lead = Lead.objects.create(
                lead_code=lead_code,
                name=name,
                mobile=mobile,
                email=email,
                hospital=target_hospital,
                campaign=campaign,
                source_category=default_cat,
                lead_source=lead_source_obj,
                stage=stage_new,
                temperature="HOT",
                deal_status="OPEN",
                admission_status="NOT_APPLIED",
                inquiry_date=inquiry_date,
                notes=notes,
                custom_data=custom_data,
                raw_source_metadata=raw_meta,
                external_lead_id=external_id,
                import_source_file=original_filename,
                import_job=job,
                created_by=request.user,
            )
            imported_count += 1

        # Automatically update Campaign start_date (earliest date) and end_date (latest date)
        if campaign and extracted_dates:
            min_d = min(extracted_dates)
            max_d = max(extracted_dates)
            if not campaign.start_date or min_d < campaign.start_date:
                campaign.start_date = min_d
            if not campaign.end_date or max_d > campaign.end_date:
                campaign.end_date = max_d
            campaign.save(update_fields=["start_date", "end_date"])

    job.imported_count = imported_count + updated_count
    job.updated_count = updated_count
    job.skipped_count = skipped_count
    job.status = ImportJob.Status.DONE
    job.completed_at = timezone.now()
    job.save()

    messages.success(
        request,
        f"✅ Leads Import Successful! {imported_count} new leads created with Campaign '{campaign.name}', "
        f"{updated_count} existing records updated, {skipped_count} duplicates skipped."
    )

    if request.user.role == User.Role.LEAD_ATTENDENT:
        return redirect("dashboard:telecaller_new_enquiries")
    return redirect("leads:lead_list")


@login_required
@user_passes_test(lambda u: u.can_import_export)
def pick_sheet(request, pk):
    job = get_object_or_404(ImportJob, pk=pk)
    sheet_name = request.POST.get("sheet_name")
    
    # Robust read
    raw = None
    for eng in ["openpyxl", "xlrd", None]:
        try:
            if eng:
                raw = pd.read_excel(job.file.path, sheet_name=sheet_name, header=None, engine=eng)
            else:
                raw = pd.read_excel(job.file.path, sheet_name=sheet_name, header=None)
            break
        except Exception:
            continue
            
    if raw is None:
        try:
            raw = pd.read_csv(job.file.path, header=None, encoding="utf-8-sig")
        except Exception:
            raw = pd.read_csv(job.file.path, header=None, encoding="latin1")

    header_row = _detect_header_row(raw)
    headers = list(raw.iloc[header_row])
    job.sheet_name = sheet_name
    job.column_mapping = {"header_row": header_row}
    job.save(update_fields=["sheet_name", "column_mapping"])

    columns = []
    date_columns = []
    for idx, h in enumerate(headers):
        if pd.isna(h):
            continue
        if _is_date_header(h):
            date_columns.append({"idx": idx, "label": str(h)})
        else:
            columns.append({"idx": idx, "label": str(h), "guess": _guess_field(h)})

    return render(request, "imports/map_columns.html", {
        "active": "import", "job": job, "columns": columns, "date_columns": date_columns,
        "target_fields": TARGET_FIELDS,
    })


def _build_mapping_from_post(request):
    mapping = {}
    for key, val in request.POST.items():
        if key.startswith("map_") and val != "IGNORE":
            idx = key.replace("map_", "")
            mapping[idx] = val
    date_cols = request.POST.getlist("date_col_idx")
    return mapping, date_cols


@login_required
@user_passes_test(lambda u: u.can_import_export)
def preview(request, pk):
    job = get_object_or_404(ImportJob, pk=pk)
    mapping, date_cols = _build_mapping_from_post(request)
    header_row = job.column_mapping.get("header_row", 0)
    job.column_mapping = {"header_row": header_row, "field_map": mapping, "date_columns": date_cols}
    job.save(update_fields=["column_mapping"])

    df = pd.read_excel(job.file.path, sheet_name=job.sheet_name, header=header_row)
    df = df.dropna(how="all")
    job.total_rows = len(df)
    job.save(update_fields=["total_rows"])

    preview_rows = []
    cols = list(df.columns)
    for _, row in df.head(10).iterrows():
        parsed = _parse_row(row, cols, mapping)
        preview_rows.append(parsed)

    return render(request, "imports/preview.html", {
        "active": "import", "job": job, "preview_rows": preview_rows, "total_rows": job.total_rows,
    })


def _parse_row(row, cols, mapping, user=None):
    data = {}
    survey_questions = []
    
    for idx_str, field in mapping.items():
        idx = int(idx_str)
        if idx >= len(cols):
            continue
        val = row.iloc[idx] if hasattr(row, "iloc") else row[cols[idx]]
        if pd.isna(val):
            val = ""
        data[field] = val

    # Gather unmapped or question-like columns into comments / survey notes
    for idx, col_name in enumerate(cols):
        idx_str = str(idx)
        field_assigned = mapping.get(idx_str, "IGNORE")
        col_str = str(col_name).strip()
        val = row.iloc[idx] if hasattr(row, "iloc") else row[col_name]
        
        if pd.isna(val) or str(val).strip() in ("", "-", "nan", "NaT"):
            continue
            
        val_str = str(val).strip()
        
        # Check if column is a survey question or non-standard custom header (e.g. Marathi/Hindi question, pregnant months, etc.)
        if field_assigned == "IGNORE":
            # Ignore technical ID columns
            col_lower = col_str.lower()
            if any(tech_id in col_lower for tech_id in ["ad_id", "adset_id", "campaign_id", "form_id", "lead_id", "hospital_id", "is_organic"]):
                continue
            if len(col_str) > 2 and not col_lower.startswith("unnamed"):
                clean_q_name = col_str.replace("_", " ").strip()
                survey_questions.append(f"[{clean_q_name}]: {val_str}")

    name = str(data.get("name", "")).strip()
    mobile, alt_mobile = cleaning.clean_phone(data.get("mobile"))
    source_cat, source_name, source_ambiguous = cleaning.normalize_source(data.get("source"))
    temperature, temp_ambiguous = cleaning.normalize_temperature(data.get("temperature"))
    inquiry_date = cleaning.parse_date(data.get("inquiry_date")) or timezone.localdate()

    # Combine explicit notes with survey questions
    base_notes = str(data.get("notes", "") or "").strip()
    all_notes_list = []
    if base_notes:
        all_notes_list.append(base_notes)
    if survey_questions:
        all_notes_list.extend(survey_questions)
    combined_notes = "\n".join(all_notes_list)

    warnings = []
    if not name or name.lower() == "nan":
        warnings.append("Missing name")
    if not mobile:
        warnings.append("Missing/invalid mobile number")

    assigned_user = _resolve_assigned_user(data.get("assigned_to"), hospital=getattr(user, 'hospital', None) if user else None)

    return {
        "name": name, "mobile": mobile, "alt_mobile": alt_mobile,
        "email": str(data.get("email", "") or "").strip(),
        "gender": str(data.get("gender", "") or "").strip(),
        "age": data.get("age", ""),
        "city": str(data.get("city", "") or "").strip(),
        "doctor": str(data.get("doctor", "") or "").strip(),
        "department": str(data.get("department", "") or "").strip(),
        "campaign_name": str(data.get("campaign", "") or "").strip(),
        "source_category": source_cat, "source_name": source_name,
        "assigned_user": assigned_user,
        "assigned_to_raw": str(data.get("assigned_to", "") or "").strip(),
        "temperature": "UNCONTACTED", "inquiry_date": inquiry_date,
        "notes": combined_notes,
        "deal_status": "OPEN",
        "admission_status": "NOT_APPLIED",
        "warnings": warnings,
    }


@login_required
@user_passes_test(lambda u: u.can_import_export or u.role in ("ADMIN", "SUPER_ADMIN", "MANAGER", "LEAD_ATTENDENT"))
def run_import(request, pk):
    from leads.models import Campaign as HospitalCampaign
    job = get_object_or_404(ImportJob, pk=pk)
    header_row = job.column_mapping.get("header_row", 0)
    mapping = job.column_mapping.get("field_map", {})
    date_cols_idx = [int(i) for i in job.column_mapping.get("date_columns", [])]
    df = None
    for eng in ["openpyxl", "xlrd", None]:
        try:
            if eng:
                df = pd.read_excel(job.file.path, sheet_name=job.sheet_name, header=header_row, engine=eng)
            else:
                df = pd.read_excel(job.file.path, sheet_name=job.sheet_name, header=header_row)
            break
        except Exception:
            continue
            
    if df is None:
        try:
            df = pd.read_csv(job.file.path, header=header_row, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(job.file.path, header=header_row, encoding="latin1")

    df = df.dropna(how="all")
    cols = list(df.columns)

    imported = updated = skipped = duplicate = invalid = 0
    default_stage = _default_stage()
    user_hospital = request.user.hospital

    for row_num, (_, row) in enumerate(df.iterrows(), start=header_row + 2):
        parsed = _parse_row(row, cols, mapping, user=request.user)
        if not parsed["name"] or not parsed["mobile"]:
            invalid += 1
            ImportErrorModel.objects.create(
                job=job, row_number=row_num,
                error_message="Missing required field(s): " + ", ".join(
                    w for w in ["Missing name", "Missing/invalid mobile number"] if w in parsed["warnings"]
                ),
                raw_row_data={str(c): str(row[c]) for c in cols[:15]},
            )
            continue

        existing = Lead.objects.filter(mobile=parsed["mobile"]).first()
        if not existing:
            digits = parsed["mobile"]
            existing = next((l for l in Lead.objects.only("id", "mobile") if Lead.clean_mobile(l.mobile) == digits), None)

        if existing and on_duplicate == "skip":
            duplicate += 1
            continue

        cat, src = _get_or_create_source(parsed["source_category"], parsed["source_name"])
        
        # Auto-register new Lead Source into Master Data & Lead Custom Field options
        from leads.models import MasterGroup, MasterItem, LeadCustomField
        if parsed["source_name"] and user_hospital:
            s_name = parsed["source_name"].strip()
            mg_src, _ = MasterGroup.objects.get_or_create(name="Lead Sources")
            MasterItem.objects.get_or_create(
                group=mg_src,
                name=s_name,
                hospital=user_hospital,
                defaults={"is_active": True}
            )
            cf_src = LeadCustomField.objects.filter(hospital=user_hospital, name="lead_source").first()
            if cf_src:
                existing_opts = [o.strip() for o in cf_src.options.split(",") if o.strip()]
                if not any(o.lower() == s_name.lower() for o in existing_opts):
                    existing_opts.append(s_name)
                    cf_src.options = ", ".join(existing_opts)
                    cf_src.save(update_fields=["options"])

        # Link or Auto-create Campaign for Hospital
        campaign_obj = None
        if parsed["campaign_name"]:
            c_name = parsed["campaign_name"].strip()
            if user_hospital:
                campaign_obj, _ = HospitalCampaign.objects.get_or_create(
                    hospital=user_hospital,
                    name=c_name,
                    defaults={"platform": parsed["source_name"] or "Meta Ads", "is_active": True}
                )
                # Auto-register new Campaign into Master Data & Lead Custom Field options
                mg_camp, _ = MasterGroup.objects.get_or_create(name="Campaigns")
                MasterItem.objects.get_or_create(
                    group=mg_camp,
                    name=c_name,
                    hospital=user_hospital,
                    defaults={"is_active": True}
                )
                cf_camp = LeadCustomField.objects.filter(hospital=user_hospital, name="campaign").first()
                if cf_camp:
                    existing_copts = [o.strip() for o in cf_camp.options.split(",") if o.strip()]
                    if not any(o.lower() == c_name.lower() for o in existing_copts):
                        existing_copts.append(c_name)
                        cf_camp.options = ", ".join(existing_copts)
                        cf_camp.save(update_fields=["options"])
            else:
                campaign_obj, _ = HospitalCampaign.objects.get_or_create(
                    name=c_name,
                    defaults={"platform": parsed["source_name"] or "Meta Ads", "is_active": True}
                )

        # Auto-register new Location/City into Master Data & Lead Custom Field options
        if parsed["city"] and user_hospital:
            city_name = parsed["city"].strip()
            mg_loc, _ = MasterGroup.objects.get_or_create(name="Locations")
            MasterItem.objects.get_or_create(
                group=mg_loc,
                name=city_name,
                hospital=user_hospital,
                defaults={"is_active": True}
            )
            cf_loc = LeadCustomField.objects.filter(hospital=user_hospital, name="location").first()
            if cf_loc:
                existing_lopts = [o.strip() for o in cf_loc.options.split(",") if o.strip()]
                if not any(o.lower() == city_name.lower() for o in existing_lopts):
                    existing_lopts.append(city_name)
                    cf_loc.options = ", ".join(existing_lopts)
                    cf_loc.save(update_fields=["options"])

        custom_data_payload = {}
        if parsed["doctor"]:
            custom_data_payload["doctor"] = parsed["doctor"]
        if parsed["department"]:
            custom_data_payload["department"] = parsed["department"]
        if parsed["age"]:
            custom_data_payload["age"] = parsed["age"]
        if parsed["gender"]:
            custom_data_payload["gender"] = parsed["gender"]

        if existing and on_duplicate == "update":
            existing.city = parsed["city"] or existing.city
            existing.email = parsed["email"] or existing.email
            if parsed.get("assigned_user"):
                existing.assigned_to = parsed["assigned_user"]
            if campaign_obj:
                existing.campaign = campaign_obj
            if parsed["notes"]:
                existing.notes = (existing.notes + "\n" + parsed["notes"]).strip()
            if custom_data_payload:
                existing.custom_data.update(custom_data_payload)
            existing.import_job = job
            existing.import_source_file = job.original_filename
            existing.save()
            lead = existing
            updated += 1
        else:
            lead = Lead.objects.create(
                name=parsed["name"], mobile=parsed["mobile"], alternate_mobile=parsed["alt_mobile"],
                email=parsed["email"], city=parsed["city"], location=parsed["city"],
                campaign=campaign_obj,
                assigned_to=parsed.get("assigned_user"),
                temperature=parsed["temperature"], stage=default_stage,
                deal_status=parsed["deal_status"], admission_status=parsed["admission_status"],
                inquiry_date=parsed["inquiry_date"], source_category=cat, lead_source=src,
                notes=parsed["notes"], created_by=request.user, hospital=user_hospital,
                custom_data=custom_data_payload,
                import_source_file=job.original_filename, import_source_sheet=job.sheet_name,
                import_source_row=row_num, import_job=job,
            )
            imported += 1

        # historical follow-up date columns -> FollowUp records
        for idx in date_cols_idx:
            if idx >= len(cols):
                continue
            comment = row.iloc[idx]
            if pd.isna(comment) or str(comment).strip() in ("", "-", "nan", "NaT"):
                continue
            fu_date = cleaning.parse_date(cols[idx]) or parsed["inquiry_date"]
            FollowUp.objects.create(
                lead=lead, followup_date=fu_date, followup_mode=FollowUpMode.OTHER,
                followup_status=FollowUpStatus.COMPLETED, comment=str(comment).strip(),
                created_by=request.user, imported_from_excel=True,
            )

    job.imported_count = imported + updated
    job.updated_count = updated
    job.duplicate_count = duplicate
    job.invalid_count = invalid
    job.status = ImportJob.Status.DONE
    job.completed_at = timezone.now()
    job.save()

    messages.success(request, f"Import complete: {imported} created, {updated} updated, {duplicate} duplicate skipped, {invalid} invalid.")
    return redirect("imports:job_detail", pk=job.pk)


@login_required
@user_passes_test(lambda u: u.can_import_export)
def job_detail(request, pk):
    job = get_object_or_404(ImportJob, pk=pk)
    return render(request, "imports/job_detail.html", {"active": "import_history", "job": job, "errors": job.errors.all()[:200]})


@login_required
@user_passes_test(lambda u: u.can_import_export)
def history(request):
    jobs = ImportJob.objects.select_related("created_by").all()
    return render(request, "imports/history.html", {"active": "import_history", "jobs": jobs})


@login_required
def export_leads(request):
    is_hospital = bool(request.user.hospital)
    is_download = request.GET.get("download") == "1"
    is_preview = request.GET.get("preview") == "1"
    
    if not is_download and not is_preview:
        from leads.models import SourceCategory, LeadSource, Campaign, Course, LeadStage
        from accounts.models import User
        
        base_leads = Lead.objects.filter(is_archived=False)
        if is_hospital:
            base_leads = base_leads.filter(hospital=request.user.hospital)
            
        used_sc_ids = base_leads.values_list("source_category_id", flat=True).distinct()
        used_stage_ids = base_leads.values_list("stage_id", flat=True).distinct()
        used_emp_ids = base_leads.values_list("assigned_to_id", flat=True).distinct()
        
        source_categories = SourceCategory.objects.filter(id__in=used_sc_ids)
        stages = LeadStage.objects.filter(id__in=used_stage_ids)
        employees = User.objects.filter(id__in=used_emp_ids)
        
        nelson_locations = []
        nelson_campaigns = []
        nelson_lead_sources = []
        nelson_deal_statuses = []
        
        if is_hospital:
            from leads.models import MasterGroup
            def get_master(name):
                grp = MasterGroup.objects.filter(name=name).first()
                if grp:
                    return grp.items.filter(hospital=request.user.hospital, is_active=True).values_list("name", flat=True)
                return []
            nelson_locations = get_master("Locations")
            nelson_campaigns = get_master("Campaigns")
            nelson_lead_sources = get_master("Lead Sources")
            nelson_deal_statuses = get_master("Deal Statuses")
            
            lead_sources = []
            campaigns = []
            distinct_cities = []
        else:
            used_ls_ids = base_leads.values_list("lead_source_id", flat=True).distinct()
            used_camp_ids = base_leads.values_list("campaign_id", flat=True).distinct()
            lead_sources = LeadSource.objects.filter(id__in=used_ls_ids)
            campaigns = Campaign.objects.filter(id__in=used_camp_ids)
            distinct_cities = sorted(list(set(base_leads.exclude(city="").values_list("city", flat=True))))
        
        # Only include courses if it's not a hospital tenant
        courses = Course.objects.all() if not is_hospital else []
        
        context = {
            "active": "export",
            "source_categories": source_categories,
            "lead_sources": lead_sources,
            "campaigns": campaigns,
            "courses": courses,
            "stages": stages,
            "employees": employees,
            "cities": distinct_cities,
            "is_hospital": is_hospital,
            "nelson_locations": nelson_locations,
            "nelson_campaigns": nelson_campaigns,
            "nelson_lead_sources": nelson_lead_sources,
            "nelson_deal_statuses": nelson_deal_statuses,
        }
        return render(request, "imports/export_leads_filter.html", context)



    from django.db.models import Q
    from leads.views import FK_FILTER_FIELDS, CHAR_FILTER_FIELDS
    leads = Lead.objects.select_related("course", "stage", "lead_source", "source_category", "assigned_to").filter(is_archived=False)
    
    if is_hospital:
        leads = leads.filter(hospital=request.user.hospital)
        
    q = request.GET.get("q", "").strip()
    if q:
        if is_hospital:
            leads = leads.filter(
                Q(lead_code__icontains=q) | Q(name__icontains=q) | Q(mobile__icontains=q)
                | Q(email__icontains=q) | Q(location__icontains=q)
                | Q(custom_data__lead_source__icontains=q) | Q(custom_data__campaign__icontains=q)
            )
        else:
            leads = leads.filter(
                Q(lead_code__icontains=q) | Q(name__icontains=q) | Q(mobile__icontains=q)
                | Q(email__icontains=q) | Q(city__icontains=q) | Q(course__name__icontains=q)
                | Q(lead_source__name__icontains=q) | Q(campaign__name__icontains=q)
            )
        
    for field in FK_FILTER_FIELDS:
        val = request.GET.get(field)
        if val:
            if is_hospital and field in ['campaign', 'lead_source']:
                leads = leads.filter(**{f"custom_data__{field}": val})
            else:
                leads = leads.filter(**{f"{field}_id": val})

    for field in CHAR_FILTER_FIELDS:
        val = request.GET.get(field)
        if val:
            leads = leads.filter(**{field: val})

    if is_hospital:
        location = request.GET.get("location")
        if location:
            leads = leads.filter(location__iexact=location)
            
        deal_status = request.GET.get("deal_status")
        if deal_status:
            leads = leads.filter(custom_data__deal_status=deal_status)
    else:
        city = request.GET.get("city")
        if city:
            leads = leads.filter(city__iexact=city)
        
        deal_status = request.GET.get("deal_status")
        if deal_status:
            leads = leads.filter(deal_status=deal_status)
            
        admission_status = request.GET.get("admission_status")
        if admission_status:
            leads = leads.filter(admission_status=admission_status)

    def _parse_date_input(val):
        if not val:
            return None
        from datetime import datetime
        for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
        return None

    date_from = _parse_date_input(request.GET.get("date_from"))
    date_to = _parse_date_input(request.GET.get("date_to"))
    if date_from:
        leads = leads.filter(inquiry_date__gte=date_from)
    if date_to:
        leads = leads.filter(inquiry_date__lte=date_to)

    def build_row(l):
        if is_hospital:
            cd = l.custom_data or {}
            doc = cd.get("doctor") or ""
            dept = cd.get("department") or ""
            loc = cd.get("location") or l.city or l.location or ""
            src = cd.get("lead_source") or (l.lead_source.name if l.lead_source else "")
            camp = cd.get("campaign") or (l.campaign.name if l.campaign else "")
            apt_st = cd.get("appointment_status") or l.display_status
            
            return {
                "Lead ID": l.lead_code,
                "Patient Name": l.name,
                "Mobile": l.mobile,
                "Email": l.email,
                "Location": loc,
                "Doctor / Consultant": doc,
                "Department": dept,
                "Lead Source": src,
                "Campaign": camp,
                "Status": l.display_status,
                "Appointment Status": apt_st,
                "Temperature": l.get_temperature_display(),
                "Inquiry Date": str(l.inquiry_date or ""),
                "Assigned To": str(l.assigned_to.get_full_name() if l.assigned_to else (cd.get("lead_attendant") or "")),
                "Next Follow-up": str(l.next_followup_date) if l.next_followup_date else "", 
                "Created At": l.created_at.strftime("%Y-%m-%d %H:%M") if l.created_at else "",
            }
        else:
            return {
                "Lead ID": l.lead_code, "Name": l.name, "Mobile": l.mobile, "Email": l.email,
                "City": l.city, "Course": str(l.course or ""),
                "Lead Source": str(l.lead_source or ""), "Campaign": str(l.campaign or ""),
                "Stage": str(l.stage), "Temperature": l.get_temperature_display(),
                "Deal Status": l.get_deal_status_display(), "Admission Status": l.get_admission_status_display(),
                "Inquiry Date": str(l.inquiry_date), "Assigned To": str(l.assigned_to or ""),
                "Next Follow-up": str(l.next_followup_date) if l.next_followup_date else "", 
                "Created At": l.created_at.strftime("%Y-%m-%d %H:%M") if l.created_at else "",
            }

    if is_preview:
        total_count = leads.count()
        preview_leads = leads.order_by("-id")[:10]
        rows = [build_row(l) for l in preview_leads]
        from django.http import JsonResponse
        return JsonResponse({"total_count": total_count, "rows": rows})

    # Download
    rows = [build_row(l) for l in leads]
    df = pd.DataFrame(rows)
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="leads_export_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    df.to_excel(response, index=False, sheet_name="Leads")
    return response


@login_required
@user_passes_test(_can_user_access_import)
def download_template(request):
    """
    Downloads Hospital-Specific Lead Import Template with Department, Doctor, Appointment Status, etc.
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.styles import Font, PatternFill, Alignment
    from leads.models import (
        HospitalDepartment, HospitalDoctor, HospitalBranch, 
        LeadSource, LeadTemperature, AppointmentStatus, Campaign as HospitalCampaign
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Hospital Leads Template"

    headers = [
        "Inquiry Date", 
        "Patient Name", 
        "Mobile", 
        "Alternate Mobile", 
        "Email", 
        "Location / City", 
        "Department", 
        "Doctor / Consultant", 
        "Campaign",
        "Lead Source", 
        "Lead Priority / Temp", 
        "Appointment Status", 
        "Notes / Medical Concern"
    ]
    
    ws.append(headers)

    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center")
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center

    sample_row = [
        "2026-08-23",
        "Adarsh Verma",
        "9617696888",
        "",
        "adarshverma753@gmail.com",
        "Nagpur",
        "Gynaecology",
        "Dr. Pradeep Patil",
        "LuxeFreeHealthCheckup_Aug21st",
        "Instagram",
        "HOT",
        "PENDING_APPROVAL",
        "Interested in consultation (4-6 months pregnant)"
    ]
    ws.append(sample_row)

    sample_font = Font(name="Calibri", size=10, italic=True, color="595959")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = sample_font

    user_hospital = request.user.hospital
    if user_hospital:
        departments = list(HospitalDepartment.objects.filter(hospital=user_hospital, is_active=True).values_list("name", flat=True))
        doctors = list(HospitalDoctor.objects.filter(hospital=user_hospital, is_active=True).values_list("name", flat=True))
        campaigns = list(HospitalCampaign.objects.filter(hospital=user_hospital, is_active=True).values_list("name", flat=True))
    else:
        departments = list(HospitalDepartment.objects.filter(is_active=True).values_list("name", flat=True))
        doctors = list(HospitalDoctor.objects.filter(is_active=True).values_list("name", flat=True))
        campaigns = list(HospitalCampaign.objects.filter(is_active=True).values_list("name", flat=True))

    sources = list(LeadSource.objects.filter(is_active=True).values_list("name", flat=True))
    if not sources:
        sources = ["Instagram", "Facebook", "Meta Ads", "Google Ads", "Website", "WhatsApp", "Walk-in"]
        
    temperatures = ["HOT", "WARM", "COLD", "UNCONTACTED"]
    appt_statuses = [choice[0] for choice in AppointmentStatus.choices]

    data_ws = wb.create_sheet(title="DropdownData")
    
    for idx, item in enumerate(departments, start=1):
        data_ws.cell(row=idx, column=1, value=item)
    for idx, item in enumerate(doctors, start=1):
        data_ws.cell(row=idx, column=2, value=item)
    for idx, item in enumerate(campaigns, start=1):
        data_ws.cell(row=idx, column=3, value=item)
    for idx, item in enumerate(sources, start=1):
        data_ws.cell(row=idx, column=4, value=item)
    for idx, item in enumerate(temperatures, start=1):
        data_ws.cell(row=idx, column=5, value=item)
    for idx, item in enumerate(appt_statuses, start=1):
        data_ws.cell(row=idx, column=6, value=item)

    data_ws.sheet_state = "hidden"

    def add_validation(col_letter, data_col_letter, count, prompt):
        if count == 0:
            return
        dv = DataValidation(
            type="list", 
            formula1=f"DropdownData!${data_col_letter}$1:${data_col_letter}${count}", 
            allow_blank=True
        )
        dv.error = 'Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = prompt
        dv.promptTitle = 'Select from list'
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}3:{col_letter}1000")

    add_validation("G", "A", len(departments), "Select a department")
    add_validation("H", "B", len(doctors), "Select a doctor")
    add_validation("I", "C", len(campaigns), "Select a campaign")
    add_validation("J", "D", len(sources), "Select a lead source")
    add_validation("K", "E", len(temperatures), "Select temperature / priority")
    add_validation("L", "F", len(appt_statuses), "Select appointment status")

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max(max_len + 3, 15)

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="nelson_hospital_leads_template.xlsx"'
    wb.save(response)
    return response


@login_required
@user_passes_test(lambda u: u.can_import_export)
def quick_import(request):
    if request.method == "POST" and "excel_file" in request.FILES:
        excel_file = request.FILES["excel_file"]
        on_duplicate = request.POST.get("on_duplicate", "skip")
        
        job = ImportJob.objects.create(
            file=excel_file,
            original_filename=excel_file.name,
            created_by=request.user,
        )
        
        real_file_path = job.file.path
        df = None
        for eng in ["openpyxl", "xlrd", None]:
            try:
                if eng:
                    df = pd.read_excel(real_file_path, sheet_name=0, engine=eng)
                else:
                    df = pd.read_excel(real_file_path, sheet_name=0)
                break
            except Exception:
                continue
                
        if df is None:
            try:
                df = pd.read_csv(real_file_path, encoding="utf-8-sig")
            except Exception:
                try:
                    df = pd.read_csv(real_file_path, encoding="latin1")
                except Exception as e:
                    job.delete()
                    messages.error(request, f"Could not read the uploaded file: {e}")
                    return redirect("imports:upload")
        
        df = df.dropna(how="all")
        df = df.dropna(how="all")
        df.columns = [str(c).strip() for c in df.columns]
        cols = list(df.columns)
        
        # Smart Dynamic Column Matcher
        def find_matching_col(aliases):
            for col in cols:
                c_clean = col.lower().replace("_", " ").strip()
                for alias in aliases:
                    if alias in c_clean or c_clean == alias:
                        return col
            return None

        col_name = find_matching_col(["patient name", "full name", "lead name", "customer name", "name", "client name"])
        col_mobile = find_matching_col(["mobile number", "phone number", "contact number", "call number", "whatsapp number", "mobile", "phone", "contact", "cell"])
        col_email = find_matching_col(["email address", "e-mail", "email", "mail"])
        col_city = find_matching_col(["city", "location", "address", "area", "town", "district"])
        col_gender = find_matching_col(["gender", "sex", "m/f"])
        col_age = find_matching_col(["age", "years", "yrs"])
        col_doctor = find_matching_col(["doctor", "dr name", "consultant", "physician", "surgeon"])
        col_dept = find_matching_col(["department", "speciality", "dept", "specialization"])
        col_campaign = find_matching_col(["campaign name", "campaign", "ad name", "ad set name"])
        col_source = find_matching_col(["lead source", "source", "platform", "publisher platform", "channel", "origin"])
        col_assigned = find_matching_col(["assigned to", "assigned", "telecaller", "executive", "attendant", "caller", "agent", "lead owner", "owner", "assignee", "counsellor"])
        col_date = find_matching_col(["created at", "created_at", "inquiry date", "lead date", "lead time", "date", "created time"])
        col_notes = find_matching_col(["remark", "comment", "issue", "note", "notes", "symptom", "problem", "query"])

        if not col_name or not col_mobile:
            job.delete()
            messages.error(
                request, 
                "Could not detect Name or Mobile column in your file. "
                "Please make sure your sheet has a column for Name (e.g. 'Patient Name', 'Name', 'Full Name') "
                "and Mobile (e.g. 'Mobile Number', 'Phone Number', 'Contact')."
            )
            return redirect("imports:upload")
            
        default_stage = _default_stage()
        user_hospital = request.user.hospital
        from leads.models import Campaign as HospitalCampaign
        
        imported = updated = skipped = duplicate = invalid = 0
        
        start_idx = 0
        if len(df) > 0:
            first_row_name = str(df.iloc[0].get(col_name, "")).strip().lower()
            first_row_mobile = str(df.iloc[0].get(col_mobile, "")).strip()
            if "rahul kumar" in first_row_name or "9876543210" in first_row_mobile:
                start_idx = 1
                
        for idx in range(start_idx, len(df)):
            row = df.iloc[idx]
            row_num = idx + 2
            
            name = str(row.get(col_name, "")).strip()
            mobile_raw = row.get(col_mobile)
            mobile, alt_mobile = cleaning.clean_phone(mobile_raw)
            
            if not name or name.lower() == "nan" or not mobile:
                invalid += 1
                continue
                
            email = str(row.get(col_email, "") or "").strip() if col_email else ""
            city = str(row.get(col_city, "") or "").strip() if col_city else ""
            gender = str(row.get(col_gender, "") or "").strip() if col_gender else ""
            age_val = row.get(col_age, "") if col_age else ""
            doctor_val = str(row.get(col_doctor, "") or "").strip() if col_doctor else ""
            dept_val = str(row.get(col_dept, "") or "").strip() if col_dept else ""
            campaign_val = str(row.get(col_campaign, "") or "").strip() if col_campaign else ""
            source_raw = str(row.get(col_source, "") or "").strip() if col_source else ""
            assigned_raw = str(row.get(col_assigned, "") or "").strip() if col_assigned else ""
            assigned_user = _resolve_assigned_user(assigned_raw, hospital=user_hospital)
            date_raw = row.get(col_date) if col_date else None
            inquiry_date = cleaning.parse_date(date_raw) or timezone.localdate()
            
            base_notes = str(row.get(col_notes, "") or "").strip() if col_notes else ""
            
            # Auto-gather survey questions from other columns (e.g. Hindi/Marathi questions, pregnant months, etc.)
            known_cols = [c for c in [col_name, col_mobile, col_email, col_city, col_gender, col_age, col_doctor, col_dept, col_campaign, col_source, col_assigned, col_date, col_notes] if c]
            survey_notes = []
            for col in cols:
                if col not in known_cols:
                    col_l = col.lower()
                    if any(t in col_l for t in ["ad_id", "adset_id", "campaign_id", "form_id", "lead_id", "hospital_id", "is_organic", "unnamed"]):
                        continue
                    v = row.get(col)
                    if pd.notna(v) and str(v).strip() not in ("", "-", "nan", "NaT"):
                        clean_col_label = col.replace("_", " ").strip()
                        survey_notes.append(f"[{clean_col_label}]: {str(v).strip()}")
                        
            all_notes = []
            if base_notes:
                all_notes.append(base_notes)
            if survey_notes:
                all_notes.extend(survey_notes)
            combined_notes = "\n".join(all_notes)
            
            source_cat, source_name, _ = cleaning.normalize_source(source_raw)
            
            existing = Lead.objects.filter(mobile=mobile).first()
            if not existing:
                existing = next((l for l in Lead.objects.only("id", "mobile") if Lead.clean_mobile(l.mobile) == mobile), None)
                
            if existing and on_duplicate == "skip":
                duplicate += 1
                continue
                
            cat, src = _get_or_create_source(source_cat, source_name)
            
            campaign_obj = None
            if campaign_val:
                if user_hospital:
                    campaign_obj, _ = HospitalCampaign.objects.get_or_create(
                        hospital=user_hospital,
                        name=campaign_val,
                        defaults={"platform": source_name or "Meta Ads", "is_active": True}
                    )
                else:
                    campaign_obj, _ = HospitalCampaign.objects.get_or_create(
                        name=campaign_val,
                        defaults={"platform": source_name or "Meta Ads", "is_active": True}
                    )
                    
            custom_data_payload = {}
            if doctor_val:
                custom_data_payload["doctor"] = doctor_val
            if dept_val:
                custom_data_payload["department"] = dept_val
            if age_val:
                custom_data_payload["age"] = age_val
            if gender:
                custom_data_payload["gender"] = gender
            if not custom_data_payload.get("priority"):
                custom_data_payload["priority"] = "Hot"
                
            if existing and on_duplicate == "update":
                existing.city = city or existing.city
                existing.email = email or existing.email
                if assigned_user:
                    existing.assigned_to = assigned_user
                if campaign_obj:
                    existing.campaign = campaign_obj
                if combined_notes:
                    existing.notes = (existing.notes + "\n" + combined_notes).strip()
                if custom_data_payload:
                    existing.custom_data.update(custom_data_payload)
                existing.import_job = job
                existing.import_source_file = job.original_filename
                existing.save()
                updated += 1
            else:
                Lead.objects.create(
                    name=name, mobile=mobile, alternate_mobile=alt_mobile,
                    email=email, city=city, location=city,
                    campaign=campaign_obj,
                    assigned_to=assigned_user,
                    temperature="HOT", stage=default_stage,
                    deal_status="OPEN", admission_status="NOT_APPLIED",
                    inquiry_date=inquiry_date, source_category=cat, lead_source=src,
                    notes=combined_notes, created_by=request.user, hospital=user_hospital,
                    custom_data=custom_data_payload,
                    import_source_file=job.original_filename, import_source_sheet="Sheet1",
                    import_source_row=row_num, import_job=job,
                )
                imported += 1

        job.imported_count = imported
        job.updated_count = updated
        job.duplicate_count = duplicate
        job.invalid_count = invalid
        job.total_rows = len(df) - start_idx
        job.status = ImportJob.Status.DONE
        job.completed_at = timezone.now()
        job.save()
        
        messages.success(
            request, 
            f"Quick import complete: {imported} created, {updated} updated, {duplicate} duplicate skipped, {invalid} invalid."
        )
        return redirect("dashboard:telecaller_new_enquiries" if request.user.role == "LEAD_ATTENDENT" else "imports:job_detail", pk=job.pk) if request.user.role != "LEAD_ATTENDENT" else redirect("dashboard:telecaller_new_enquiries")
        
    return redirect("imports:upload")


@login_required
@user_passes_test(lambda u: u.can_import_export)
def delete_import(request, pk):
    if request.method == "POST":
        job = get_object_or_404(ImportJob, pk=pk)
        from leads.models import Lead
        leads = Lead.objects.filter(import_job=job)
        leads_count = leads.count()
        leads.delete()
        if job.file:
            try:
                job.file.delete(save=False)
            except Exception:
                pass
        job.delete()
        messages.success(request, f"Import history item and its {leads_count} associated leads have been deleted successfully.")
    return redirect("imports:history")
